#!/usr/bin/env python3
"""
Extract recipient emails from the Programadores Cariocas matriculados xlsx.

Data sources (sheets in the xlsx):
  - "Alunos Generation (Inscritos)": email in column A directly.
  - "Alunos Generation (Alunos mês 1[2])": email in column A directly.
  - "Alunos Senac (Inscritos)": CPF only (column A). No email.
  - "Alunos Senac (Alunos mês 12)": CPF only (column A). No email.
  - "Senac_Procv": lookup table — CPF in column A, email in column H.
    We join the Senac student sheets to this by CPF to recover emails.

Email normalization:
  Many entries have "@gmailcom" (missing dot) instead of "@gmail.com".
  Same for hotmail, outlook, yahoo, etc. This script fixes those.

Output:
  - recipients_all.local.txt     (deduplicated, all sources — gitignored)
  - recipients_generation.local.txt (Generation sources only)
  - recipients_senac.local.txt   (Senac sources only)
  - stats.md                     (aggregate counts — safe to commit)

Usage:
  python3 extract_emails.py <path_to_xlsx>
"""

import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

KNOWN_DOMAINS = [
    "gmail.com", "hotmail.com", "outlook.com", "yahoo.com",
    "yahoo.com.br", "live.com", "icloud.com", "uol.com.br",
    "bol.com.br", "terra.com.br", "globo.com", "ig.com.br",
    "protonmail.com", "msn.com",
    "hotmail.com.br", "outlook.com.br", "live.com.br",
    "gmail.com.br",
]

DOMAIN_FIXES = {}
for d in KNOWN_DOMAINS:
    broken = d.replace(".", "")
    if broken != d:
        DOMAIN_FIXES[broken] = d
    no_dot_before_br = d.replace(".com.br", "combr")
    if no_dot_before_br != d:
        DOMAIN_FIXES[no_dot_before_br] = d


def normalize_email(raw):
    if not raw or not isinstance(raw, str):
        return None
    email = raw.strip().lower()
    email = email.replace(" ", "")
    if "@" not in email:
        return None
    local, domain = email.rsplit("@", 1)
    if not local or not domain:
        return None
    if domain in DOMAIN_FIXES:
        domain = DOMAIN_FIXES[domain]
    elif "." not in domain:
        for known in KNOWN_DOMAINS:
            if known.replace(".", "") == domain:
                domain = known
                break
    email = f"{local}@{domain}"
    if not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
        return None
    return email


def normalize_cpf(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    s = re.sub(r'[.\-/\s]', '', s)
    if not s.isdigit():
        return None
    return s.zfill(11)


def main():
    if len(sys.argv) < 2:
        sys.exit(f"Usage: {sys.argv[0]} <path_to_xlsx>")

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    out_dir = Path(__file__).resolve().parent

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    generation_emails = set()
    senac_emails = set()
    malformed = []
    cpf_no_email = []

    # --- Generation sheets: email is in column A ---
    gen_sheets = [s for s in wb.sheetnames if "generation" in s.lower() and "procv" not in s.lower()]
    for sheet_name in gen_sheets:
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[0] if row else None
            email = normalize_email(str(raw) if raw else None)
            if email:
                generation_emails.add(email)
                count += 1
            elif raw and str(raw).strip():
                malformed.append((sheet_name, str(raw).strip()))
        print(f"  {sheet_name}: {count} valid emails extracted")

    # --- Senac_Procv: build CPF -> email lookup ---
    cpf_to_email = {}
    procv_sheet = "Senac_Procv"
    if procv_sheet in wb.sheetnames:
        ws = wb[procv_sheet]
        header = None
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                header = [str(c).lower() if c else "" for c in row]
                continue
            cpf_raw = row[0] if row else None
            cpf = normalize_cpf(cpf_raw)
            if not cpf:
                continue
            email_raw = row[7] if len(row) > 7 else None
            email = normalize_email(str(email_raw) if email_raw else None)
            if email and cpf:
                cpf_to_email[cpf] = email
        print(f"\n  {procv_sheet}: {len(cpf_to_email)} CPF->email mappings loaded")
    else:
        print(f"\n  WARNING: {procv_sheet} sheet not found!")

    # --- Senac student sheets: CPF in column A, join to Procv ---
    senac_sheets = [s for s in wb.sheetnames if "senac" in s.lower()
                    and "procv" not in s.lower()]
    for sheet_name in senac_sheets:
        ws = wb[sheet_name]
        found = 0
        missed = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cpf_raw = row[0] if row else None
            cpf = normalize_cpf(cpf_raw)
            if not cpf:
                continue
            email = cpf_to_email.get(cpf)
            if email:
                senac_emails.add(email)
                found += 1
            else:
                missed += 1
                cpf_no_email.append((sheet_name, cpf))
        print(f"  {sheet_name}: {found} emails via CPF join, {missed} CPFs with no email in Procv")

    wb.close()

    # --- Dedup and write ---
    all_emails = generation_emails | senac_emails
    overlap = generation_emails & senac_emails

    print(f"\n--- Summary ---")
    print(f"  Generation (unique): {len(generation_emails)}")
    print(f"  Senac (unique):      {len(senac_emails)}")
    print(f"  Overlap:             {len(overlap)}")
    print(f"  Total (deduplicated):{len(all_emails)}")
    print(f"  Malformed entries:   {len(malformed)}")
    print(f"  Senac CPFs w/o email:{len(cpf_no_email)}")

    def write_list(path, emails):
        sorted_emails = sorted(emails)
        path.write_text("\n".join(sorted_emails) + "\n", encoding="utf-8")
        print(f"  Wrote {len(sorted_emails)} emails to {path.name}")

    write_list(out_dir / "recipients_all.local.txt", all_emails)
    write_list(out_dir / "recipients_generation.local.txt", generation_emails)
    write_list(out_dir / "recipients_senac.local.txt", senac_emails)

    # --- Stats (safe to commit) ---
    stats = [
        "# Email Extraction Stats",
        "",
        f"Source file: Lista_Matriculados_Definitica_por_Ra_a_Genero.xlsx",
        f"Extraction date: {__import__('datetime').date.today()}",
        "",
        "## Counts",
        "",
        f"| Source | Unique emails |",
        f"|--------|--------------|",
        f"| Generation sheets | {len(generation_emails)} |",
        f"| Senac (via CPF join to Senac_Procv) | {len(senac_emails)} |",
        f"| Overlap (in both) | {len(overlap)} |",
        f"| **Total (deduplicated)** | **{len(all_emails)}** |",
        "",
        "## Data quality",
        "",
        f"- Malformed entries skipped: {len(malformed)}",
        f"- Senac CPFs with no email in Procv: {len(cpf_no_email)}",
        "",
        "## How emails were extracted",
        "",
        "- **Generation sheets** (`Alunos Generation (Inscritos)` and",
        "  `Alunos Generation (Alunos mês 1[2])`): email is directly in",
        "  column A. Normalized domain typos (`@gmailcom` -> `@gmail.com`).",
        "- **Senac sheets** (`Alunos Senac (Inscritos)` and `Alunos Senac",
        "  (Alunos mês 12)`): only have CPF (column A), no email. Joined",
        "  each CPF to the `Senac_Procv` sheet (3663-row lookup table)",
        "  where CPF is column A and email is column H.",
        "- Deduplicated across all sources (case-insensitive).",
        "",
        "## Domain normalization applied",
        "",
        "Many entries had missing dots in domains:",
        "- `@gmailcom` -> `@gmail.com`",
        "- `@hotmailcom` -> `@hotmail.com`",
        "- `@outlookcom` -> `@outlook.com`",
        "- `@yahoocom` -> `@yahoo.com`",
        "- etc.",
    ]

    if malformed:
        stats.append("")
        stats.append("## Malformed entries (could not fix)")
        stats.append("")
        for sheet, val in malformed[:20]:
            stats.append(f"- `{sheet}`: `{val[:50]}`")
        if len(malformed) > 20:
            stats.append(f"- ... and {len(malformed) - 20} more")

    stats_path = out_dir / "stats.md"
    stats_path.write_text("\n".join(stats) + "\n", encoding="utf-8")
    print(f"  Wrote stats to {stats_path.name}")


if __name__ == "__main__":
    main()
